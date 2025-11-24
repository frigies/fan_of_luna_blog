import os
import pandas as pd
import psycopg2
from openpyxl import load_workbook
import re
import requests
from urllib.parse import urlparse

# Настройки для requests
SESSION = requests.Session()
SESSION.headers.update({
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
})

# Функция для получения финального URL после редиректов
def get_final_url(url, timeout=5):
    """
    Получает финальный URL после всех редиректов
    """
    if not url or not url.startswith(('http://', 'https://')):
        return url
    
    try:
        response = SESSION.head(url, allow_redirects=True, timeout=timeout)
        return response.url
    except requests.RequestException as e:
        print(f"Ошибка при проверке URL {url}: {e}")
        return None

# Функция для обработки специальных случаев редиректов
def process_special_redirects(url):
    """
    Обрабатывает известные сайты-редиректы
    """
    if not url:
        return url
    
    parsed = urlparse(url)
    domain = parsed.netloc.lower()
    
    # Обрабатываем bitcoin-vps.com
    if 'bitcoin-vps.com' in domain:
        try:
            print(f"Обрабатываем bitcoin-vps.com редирект: {url}")
            response = SESSION.get(url, timeout=10, allow_redirects=False)
            if response.status_code in [301, 302] and 'Location' in response.headers:
                final_url = response.headers['Location']
                print(f"Найден редирект: {url} -> {final_url}")
                return final_url
        except requests.RequestException as e:
            print(f"Ошибка при обработке bitcoin-vps.com: {e}")
    
    return url

# Читаем XLSX файл
def read_xlsx_with_hyperlinks(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active
    
    headers = []
    for cell in sheet[1]:
        headers.append(cell.value)
    
    rows = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        if all(cell.value is None for cell in row):
            continue
            
        row_data = []
        for cell in row:
            if cell.hyperlink:
                row_data.append({
                    'text': cell.value,
                    'url': cell.hyperlink.target
                })
            else:
                row_data.append(cell.value)
        
        rows.append(row_data)
    
    df = pd.DataFrame(rows, columns=headers)
    print(f"Прочитано строк: {len(rows)}")
    return df

# Функция для обработки гиперссылок с редиректами
def parse_hyperlink_with_redirects(cell):
    if cell is None:
        return "", None
    
    # Если это словарь с гиперссылкой из openpyxl
    if isinstance(cell, dict) and 'text' in cell and 'url' in cell:
        text = cell['text'] if cell['text'] is not None else ""
        url = cell['url'] if cell['url'] is not None else ""
        
        # Обрабатываем редиректы для bitcoin-vps.com
        if url and 'bitcoin-vps.com' in url:
            final_url = process_special_redirects(url)
            if final_url and final_url != url:
                print(f"Обновлена ссылка: {url} -> {final_url}")
                url = final_url
        
        return text, url
    
    # Если это строка
    cell_str = str(cell) if cell is not None else ""
    
    # Telegram бот
    if cell_str.startswith('@'):
        username = cell_str
        url = f"https://t.me/{username[1:]}"
        return username, url
    
    # Проверяем, есть ли URL в тексте
    url_pattern = r'(https?://[^\s]+)'
    url_match = re.search(url_pattern, cell_str)
    if url_match:
        url = url_match.group(1)
        
        # Обрабатываем редиректы
        if 'bitcoin-vps.com' in url:
            final_url = process_special_redirects(url)
            if final_url and final_url != url:
                print(f"Обновлена ссылка из текста: {url} -> {final_url}")
                url = final_url
        
        name = re.sub(url_pattern, '', cell_str).strip()
        return name, url
    
    # Просто текст - проверяем, похоже ли на домен
    if re.match(r'^[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', cell_str):
        url = f"https://{cell_str}"
        
        # Проверяем редиректы для доменов
        if 'bitcoin-vps.com' in url:
            final_url = process_special_redirects(url)
            if final_url and final_url != url:
                print(f"Обновлена ссылка домена: {url} -> {final_url}")
                url = final_url
        
        return cell_str, url
    
    # Просто текст без ссылки
    return cell_str, None

# Улучшенная функция для обработки цены
def parse_price(price_cell):
    if price_cell is None:
        return None
    
    price_str = str(price_cell).strip()
    
    # Если пустая строка или не число, возвращаем NULL
    if not price_str or price_str == '':
        return None
    
    # Удаляем символ $ и пробелы
    price_str = price_str.replace('$', '').replace(' ', '').replace(',', '.').strip()
    
    # Проверяем, является ли числом
    try:
        # Пробуем преобразовать в float
        price_float = float(price_str)
        return price_float
    except (ValueError, TypeError):
        print(f"Невозможно преобразовать цену: '{price_cell}' -> '{price_str}'")
        return None

# Основной скрипт
print("Чтение XLSX файла с гиперссылками...")
df = read_xlsx_with_hyperlinks('hostings.xlsx')

print(f"Всего строк в DataFrame: {len(df)}")

# Обрабатываем гиперссылки с редиректами
print("\nОбработка гиперссылок с редиректами...")
df[['hosting_name', 'url']] = df['Хостинг'].apply(
    lambda x: pd.Series(parse_hyperlink_with_redirects(x))
)

# Обрабатываем цену
print("\nОбработка цен...")
df['min_price_in_dollars'] = df['Минимальная цена'].apply(parse_price)

# Проверяем уникальность имен хостингов
duplicate_names = df[df.duplicated('hosting_name', keep=False)]
if not duplicate_names.empty:
    print(f"\nПредупреждение: найдены дубликаты имен хостингов:")
    for name in duplicate_names['hosting_name'].unique():
        duplicates = duplicate_names[duplicate_names['hosting_name'] == name]
        print(f"  '{name}': {len(duplicates)} раз")

# Переименовываем столбцы
column_mapping = {
    'Статус': 'status',
    'Значение риска proxycheck.io (APIv3)': 'risk',
    'Преимущества': 'advantages', 
    'Недостатки': 'disadvantages',
    'Расположение хостинга': 'hosting_location',
    'Расположение серверов': 'servers_location'
}

for old_col, new_col in column_mapping.items():
    if old_col in df.columns:
        df[new_col] = df[old_col]
    else:
        df[new_col] = None

# Заменяем NaN на None
df = df.where(pd.notnull(df), None)

# Статистика
print(f"\nСтатистика обработки:")
print(f"Всего записей: {len(df)}")
print(f"Ссылки: {df['url'].notna().sum()}")
print(f"Без ссылок: {df['url'].isna().sum()}")
print(f"Telegram ботов: {df['hosting_name'].str.startswith('@').fillna(False).sum()}")
print(f"Цены указаны: {df['min_price_in_dollars'].notna().sum()}")
print(f"Цены отсутствуют: {df['min_price_in_dollars'].isna().sum()}")

# Подключаемся к PostgreSQL
db_host = os.getenv('BLOG_DB_HOST')
db_name = os.getenv('BLOG_DB_NAME')
db_user = os.getenv('BLOG_DB_USER')
db_password = os.getenv('BLOG_DB_PASSWORD')

try:
    conn = psycopg2.connect(
        dbname=db_name, 
        user=db_user, 
        password=db_password, 
        host=db_host
    )
    cur = conn.cursor()

    # Вставляем данные с обработкой дубликатов
    inserted_count = 0
    error_count = 0
    duplicate_count = 0
    
    for idx, row in df.iterrows():
        try:
            hosting_name = str(row['hosting_name'])[:500] if row['hosting_name'] else None
            
            # Проверяем дубликаты
            cur.execute("SELECT hosting_id FROM hosting WHERE hosting_name = %s", (hosting_name,))
            if cur.fetchone():
                print(f"Пропущен дубликат: {hosting_name}")
                duplicate_count += 1
                continue
            
            # Подготавливаем данные
            url = str(row['url'])[:1000] if row['url'] else None
            status = str(row.get('status'))[:500] if row.get('status') else None
            
            risk_value = None
            if pd.notna(row.get('risk')):
                risk_str = str(row.get('risk')).strip()
                if risk_str.isdigit():
                    risk_value = int(risk_str)
            
            advantages = str(row.get('advantages'))[:2000] if row.get('advantages') else None
            disadvantages = str(row.get('disadvantages'))[:2000] if row.get('disadvantages') else None
            hosting_location = str(row.get('hosting_location'))[:500] if row.get('hosting_location') else None
            servers_location = str(row.get('servers_location'))[:500] if row.get('servers_location') else None
            min_price = row.get('min_price_in_dollars')
            
            cur.execute("""
                INSERT INTO hosting (
                    hosting_name, url, status, risk, advantages, 
                    disadvantages, hosting_location, servers_location, min_price_in_dollars
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING hosting_id
            """, (
                hosting_name, url, status, risk_value, advantages,
                disadvantages, hosting_location, servers_location, min_price
            ))
            
            hosting_id = cur.fetchone()[0]
            inserted_count += 1
            
            # Присваиваем категории
            categories_to_assign = [2]  # Настройте по необходимости
            for cat_id in categories_to_assign:
                try:
                    cur.execute(
                        "INSERT INTO hosting_category (hosting_id, category_id) VALUES (%s, %s)",
                        (hosting_id, cat_id)
                    )
                except Exception as e:
                    print(f"Ошибка при добавлении категории {cat_id} для hosting_id {hosting_id}: {e}")
                    
        except psycopg2.IntegrityError as e:
            if 'unique_hosting_name' in str(e):
                duplicate_count += 1
                print(f"Дубликат (constraint): {hosting_name}")
            else:
                error_count += 1
                print(f"Ошибка Integrity при вставке {hosting_name}: {e}")
        except Exception as e:
            error_count += 1
            print(f"Ошибка при вставке строки {idx+2}: {hosting_name} - {e}")

    conn.commit()
    print(f"\nИмпорт завершён!")
    print(f"Успешно вставлено: {inserted_count}")
    print(f"Дубликатов пропущено: {duplicate_count}")
    print(f"Ошибок: {error_count}")

except Exception as e:
    print(f"Ошибка при работе с базой данных: {e}")
    if conn:
        conn.rollback()
finally:
    if cur:
        cur.close()
    if conn:
        conn.close()