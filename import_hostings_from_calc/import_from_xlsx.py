import os
import pandas as pd
import psycopg2
from openpyxl import load_workbook
import re

# Читаем XLSX файл с сохранением гиперссылок
def read_xlsx_with_hyperlinks(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active
    
    # Получаем заголовки
    headers = []
    for cell in sheet[1]:
        headers.append(cell.value)
    
    # Собираем данные
    rows = []
    for row in sheet.iter_rows(min_row=2):  # пропускаем заголовки
        row_data = []
        for cell in row:
            # Проверяем есть ли гиперссылка
            if cell.hyperlink:
                # Сохраняем и текст и ссылку
                row_data.append({
                    'text': cell.value,
                    'url': cell.hyperlink.target
                })
            else:
                # Просто значение ячейки
                row_data.append(cell.value)
        rows.append(row_data)
    
    # Создаем DataFrame
    df = pd.DataFrame(rows, columns=headers)
    return df

# Читаем XLSX файл
print("Чтение XLSX файла с гиперссылками...")
df = read_xlsx_with_hyperlinks('hostings.xlsx')

print("Первые строки из XLSX файла:")
print(df.head(10))
print("\nСтолбцы:")
print(df.columns.tolist())

# Функция для обработки гиперссылок
def parse_hyperlink(cell):
    if cell is None:
        return "", None
    
    # Если это словарь с гиперссылкой из openpyxl
    if isinstance(cell, dict) and 'text' in cell and 'url' in cell:
        return cell['text'], cell['url']
    
    # Если это строка
    cell_str = str(cell)
    
    # Telegram бот
    if cell_str.startswith('@'):
        username = cell_str
        url = f"https://t.me/{username[1:]}"
        return username, url
    
    # Проверяем, есть ли URL в тексте (ручной парсинг)
    url_pattern = r'(https?://[^\s]+)'
    url_match = re.search(url_pattern, cell_str)
    if url_match:
        url = url_match.group(1)
        name = re.sub(url_pattern, '', cell_str).strip()
        return name, url
    
    # Просто текст - проверяем, похоже ли на домен
    if re.match(r'^[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', cell_str):
        return cell_str, f"https://{cell_str}"
    
    # Просто текст без ссылки
    return cell_str, None

# Функция для обработки цены
def parse_price(price_cell):
    if price_cell is None:
        return None
    
    price_str = str(price_cell)
    
    # Удаляем символ $ и пробелы, преобразуем в число
    price_str = price_str.replace('$', '').replace(' ', '').strip()
    
    try:
        return float(price_str) if price_str else None
    except (ValueError, TypeError):
        return None

# Диагностика формата данных
print("\nДиагностика формата данных в столбце 'Хостинг':")
for i, cell in enumerate(df['Хостинг'].head(15)):
    cell_type = type(cell)
    if isinstance(cell, dict):
        print(f"{i}: text='{cell['text']}', url='{cell['url']}' (тип: dict)")
    else:
        print(f"{i}: {repr(cell)} (тип: {cell_type})")

# Обрабатываем гиперссылки
print("\nОбработка гиперссылок...")
df[['hosting_name', 'url']] = df['Хостинг'].apply(
    lambda x: pd.Series(parse_hyperlink(x))
)

# Обрабатываем остальные столбцы
df['min_price_in_dollars'] = df['Минимальная цена'].apply(parse_price)

# Переименовываем столбцы
column_mapping = {
    'Статус': 'status',
    'Значение риска proxycheck.io (APIv3)': 'risk',
    'Преимущества': 'advantages', 
    'Недостатки': 'disadvantages',
    'Расположение хостинга': 'hosting_location',
    'Расположение серверов': 'servers_location'
}

for old_col, new_col in column_mapping.items():
    if old_col in df.columns:
        df[new_col] = df[old_col]
    else:
        df[new_col] = None

# Заменяем NaN на None
df = df.where(pd.notnull(df), None)

# Выводим статистику
print(f"\nСтатистика обработки:")
print(f"Всего записей: {len(df)}")
print(f"С ссылками: {df['url'].notna().sum()}")
print(f"Без ссылок: {df['url'].isna().sum()}")
print(f"Telegram ботов: {df['hosting_name'].str.startswith('@').fillna(False).sum()}")

print("\nПримеры обработанных данных:")
for i, row in df[['hosting_name', 'url', 'min_price_in_dollars']].head(15).iterrows():
    print(f"{i}: name='{row['hosting_name']}', url={row['url']}")

# Покажем записи, где есть ссылки
has_links = df[df['url'].notna()]
print(f"\nНайдено записей со ссылками: {len(has_links)}")
for i, row in has_links[['hosting_name', 'url']].head(20).iterrows():
    print(f"  {row['hosting_name']} -> {row['url']}")

# Подключаемся к PostgreSQL
db_host = os.getenv('BLOG_DB_HOST')
db_name = os.getenv('BLOG_DB_NAME')
db_user = os.getenv('BLOG_DB_USER')
db_password = os.getenv('BLOG_DB_PASSWORD')

try:
    conn = psycopg2.connect(
        dbname=db_name, 
        user=db_user, 
        password=db_password, 
        host=db_host
    )
    cur = conn.cursor()

    # Вставляем данные
    inserted_count = 0
    for _, row in df.iterrows():
        try:
            cur.execute("""
                INSERT INTO hosting (
                    hosting_name, url, status, risk, advantages, 
                    disadvantages, hosting_location, servers_location, min_price_in_dollars
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING hosting_id
            """, (
                str(row['hosting_name'])[:500] if row['hosting_name'] else None,
                str(row['url'])[:1000] if row['url'] else None,
                str(row.get('status'))[:500] if row.get('status') else None,
                int(row.get('risk')) if pd.notna(row.get('risk')) and str(row.get('risk')).strip().isdigit() else None,
                str(row.get('advantages'))[:2000] if row.get('advantages') else None,
                str(row.get('disadvantages'))[:2000] if row.get('disadvantages') else None,
                str(row.get('hosting_location'))[:500] if row.get('hosting_location') else None,
                str(row.get('servers_location'))[:500] if row.get('servers_location') else None,
                row.get('min_price_in_dollars')
            ))
            
            hosting_id = cur.fetchone()[0]
            inserted_count += 1
            
            # Присваиваем категории
            categories_to_assign = [2]  # Настройте по необходимости
            for cat_id in categories_to_assign:
                try:
                    cur.execute(
                        "INSERT INTO hosting_category (hosting_id, category_id) VALUES (%s, %s)",
                        (hosting_id, cat_id)
                    )
                except Exception as e:
                    print(f"Ошибка при добавлении категории {cat_id} для hosting_id {hosting_id}: {e}")
                    
        except Exception as e:
            print(f"Ошибка при вставке записи {row['hosting_name']}: {e}")
            continue

    conn.commit()
    print(f"\nИмпорт завершён успешно! Вставлено записей: {inserted_count}/{len(df)}")

except Exception as e:
    print(f"Ошибка при работе с базой данных: {e}")
    if conn:
        conn.rollback()
finally:
    if cur:
        cur.close()
    if conn:
        conn.close()